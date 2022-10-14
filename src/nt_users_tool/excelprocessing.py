from openpyxl import Workbook, load_workbook
from datetime import date

from constants import *


def create_results_workbook(excel_desired_path: str, sheets_names_list: list):
    """Creates the output excel file according to arguments

    Args:
        excel_desired_path (str): Path where the output file should be located
        sheets_names_list (list): Names of the sheets inside the file
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in sheets_names_list:
        workbook.create_sheet(sheet_name)
    workbook.save(excel_desired_path)
    workbook.close()


def fill_all_sheets(excel_file_name: str, user_info: list):
    results = load_workbook(excel_file_name, read_only=False)

    expired_users_sheet = results[SHEET_EXPIRED]
    expiring_soon_users_sheet = results[SHEET_EXPIRES_SOON]
    users_sheet = results[SHEET_ALL_USERS]

    today = date.today()
    month, year = str(today.month), str(today.year)

    all_user_index = 1
    expired_index = 1
    expiring_index = 1

    for user in user_info:
        fill_one_row(users_sheet, all_user_index, COLUMNS_LIST, user)
        all_user_index += 1

        user_day, user_month, user_year = user[EXPIRATION_DATE].split(SLASH)
        if user_year < year:
            fill_one_row(expired_users_sheet, expired_index, COLUMNS_LIST, user)
            expired_index += 1

    results.save(excel_file_name)
    results.close()


def fill_one_row(worksheet, row_number: int, columns: list, user_info: dict):
    """_summary_

    :param worksheet: _description_
    :type worksheet: _type_
    :param row_number: _description_
    :param columns: _description_
    :param user_info: _description_
    """

    keys = list(user_info.keys())
    for i, column in enumerate(columns):
        print(column + str(row_number), keys[i], user_info[keys[i]])
        worksheet[column + str(row_number)] = user_info[keys[i]]


def nt_user_list(excel_file_name: str):
    """Returns the list of nt_user IDs , reading them from the excel file.

    Args:
        excel_file_name (str): Excel Input filepath

    Returns:
        list: List of nt_user IDs
    """
    workbook = load_workbook(excel_file_name, read_only=True)
    worksheet = workbook.active
    nt_user_id_list = []
    for row in worksheet.rows:
        for cell in row:
            if cell.value != None:
                nt_user_id_list.append(cell.value)
    workbook.close()
    return nt_user_id_list
