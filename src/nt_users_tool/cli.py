import logging
import logging.config
import os
import sys
from datetime import date
from sys import exit
from time import perf_counter

import click
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from nt_users_tool.constants import DEFAULT_INPUT_FILE, SHEET_INPUT
from nt_users_tool.excel_processing import create_results_sheets, fill_all_sheets, read_nt_users
from nt_users_tool.net_commands import extract_all_nt_user_info, get_all_nt_user_string


@click.command()
@click.option(
    "--filename",
    type=click.Path(exists=True),
    default=".",
    help="Path and name to .xlsx file",
)
def main(filename) -> int:
    """The main functions takes an excel file with a list of nt_users as input.
    It runs the net command /domain for each nt_user and generates multiple sheets to portray the current state of each user.
    It also creates a table on the last sheet to easily sort through.
    Everything is saved on the same input document.

    :return: 1 if the script executed successfully, 1 if an error was raised.
    """

    # Paths to the config folder, depend if python code or executable file (bundled app)
    #   When a bundled app starts up, the bootloader sets the sys.frozen attribute
    #   and stores the absolute path to the bundle folder in sys._MEIPASS.
    is_pyinstaller_bundle = getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS")
    if is_pyinstaller_bundle:
        # this give path to "config" folder in the bundled app
        this_script_dir = os.path.join(sys._MEIPASS)
        default_log_cfg_file = os.path.join(this_script_dir, "config", "log.ini")
    else:
        # this give the path to the workspace directory (python project folder)
        this_script_dir = os.path.dirname(os.path.realpath(__file__))
        default_log_cfg_file = os.path.join(this_script_dir, "..", "..", "config", "log.ini")

    log_config_path = os.environ.get("LOG_CONFIG_INI", default_log_cfg_file)

    # import logging config from log.ini
    logging.config.fileConfig(log_config_path)
    logger = logging.getLogger("sLogger")
    logger.info("Starting nt_user verification.")
    logger.info("nt_users_tool started !")

    # Cath the path of the xlsx file
    if filename == ".":
        # Ask for path and name of input file
        filename = (
            input(f"What's the PATH and the name of your file ? (PRESS ENTER if default {DEFAULT_INPUT_FILE}) : ")
            or DEFAULT_INPUT_FILE
        )
    try:
        wb_input = load_workbook(filename, read_only=False)
    except InvalidFileException:
        logger.warning(f"{filename} is not a valid excel file. Supported formats are: .xlsx,.xlsm,.xltx,.xltm")
        return 1
    except FileNotFoundError:
        logger.warning(f"{filename} cannot be found in current directory. Make sure it is present.")
        return 1
    except PermissionError:
        logger.warning(f"{filename} cannot be opened. Make sure it is not open in another program.")
        return 1
    logger.info(f"Opening {filename}.")

    # Try to read the list of NT users in the input sheet
    try:
        ws = wb_input[SHEET_INPUT]
    except KeyError:
        logger.warning(
            f"Could not open the worksheet named {SHEET_INPUT} inside {filename}. Make sure the nt users are listed in this sheet."
        )
        return 1

    logger.info(f"Reading worksheet {SHEET_INPUT}.")
    nt_users = read_nt_users(ws)

    logger.info("Gathering information from the network...")
    start = perf_counter()
    net_command_responses = get_all_nt_user_string(nt_users)
    end = perf_counter()
    logger.info(f"Information gathered in {end-start:.2f} seconds.")

    logger.info("Filtering.")
    nt_user_info_list = extract_all_nt_user_info(net_command_responses)

    date_today = date.today().strftime("%Y-%m-%d")
    output_file_name = f"nt_users_sorted_{date_today}.xlsx"
    logger.info(f"Write results to {output_file_name}.")
    wb_output = Workbook()
    create_results_sheets(wb_output)
    fill_all_sheets(wb_output, nt_user_info_list)
    del wb_output["Sheet"]  # remove intial sheet "Sheet" if exist

    # Save data organized in an new Excel workbook
    try:
        wb_output.save(filename=output_file_name)
    except PermissionError:
        logger.warning(f"Could not save {output_file_name}. Make sure it is not open in another program.")
        return 1

    wb_output.close()
    wb_input.close()
    logger.info(f"Done. {output_file_name} saved with changes.")
    logger.info("nt_users_tool finished !")
    return 0


if __name__ == "__main__":
    exit(main())
