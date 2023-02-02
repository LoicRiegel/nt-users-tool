from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from nt_users_tool.constants import COLUMNS_LIST
from nt_users_tool.constants import Sheets
from nt_users_tool.constants import Sheets as sh
from nt_users_tool.constants import SheetSettingsInt as shi
from nt_users_tool.constants import SheetSettingsStr as shs
from nt_users_tool.constants import Tables as tb
from nt_users_tool.nt_user_info import NTUserInfo, NTUserStatus, evaluate_user_status

# Processing input


def read_nt_users(worksheet: Worksheet) -> List[str]:
    """Generates a list of nt_user found in the worksheet.

    :param worksheet: The worksheet with nt_users inside.
    :return: A list of nt_user (string).
    """
    nt_user_id_list = []
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                nt_user_id_list.append(cell.value)
    return nt_user_id_list


# Generating output


def create_results_sheets(workbook: Workbook):
    """Creates the sheets inside the given workbook according to parameter.

    :param workbook: The workbook in which to add the sheets.
    """
    for sheet_name in Sheets:
        workbook.create_sheet(sheet_name.value)
        workbook[sheet_name.value].column_dimensions[shs.NAME_COLUMN.value].width = shi.COLUMN_WIDTH.value
        workbook[sheet_name.value].column_dimensions[shs.NT_USER_COLUMN.value].width = shi.COLUMN_WIDTH.value / 2
        workbook[sheet_name.value].column_dimensions[shs.EXPIRATION_DATE_COLUMN.value].width = shi.COLUMN_WIDTH.value


def fill_one_row(worksheet: Worksheet, row_number: int, columns: List[str], nt_user_info: NTUserInfo):
    """Fills row_number of columns on the given worksheet, with the info in nt_user_info.

    :param worksheet: The worksheet that needs to be changed.
    :param row_number: the row at which the changes happen.
    :param columns: the range of columns we write on.
    :param user_info: The information that we are writing.
    """
    for i, column in enumerate(columns):
        worksheet[column + str(row_number)] = nt_user_info[i]


def fill_all_sheets(workbook: Workbook, nt_user_info_list: List[NTUserInfo]):
    """Fills all sheets for the given workbook and nt_user info list (via fill_one_row).

    :param workbook: The workbook which has all sheets to be modified.
    :param nt_user_info_list: The list of NTUserInfo objects containing information.
    """
    all_users_index = 2
    error_name_index = 2
    expired_index = 2
    expiring_15_index = 2
    expiring_30_index = 2
    expiring_60_index = 2

    for user_info in nt_user_info_list:
        fill_one_row(workbook[sh.SHEET_ALL_USERS.value], all_users_index, COLUMNS_LIST, user_info)
        all_users_index += 1
        user_status = evaluate_user_status(user_info)
        if user_status == NTUserStatus.INVALID_NAME:
            fill_one_row(
                workbook[sh.SHEET_ERROR_NAME_USER.value],
                error_name_index,
                COLUMNS_LIST,
                user_info,
            )
            error_name_index += 1
        if user_status == NTUserStatus.EXPIRED:
            fill_one_row(workbook[sh.SHEET_EXPIRED.value], expired_index, COLUMNS_LIST, user_info)
            expired_index += 1
        elif user_status == NTUserStatus.EXPIRING_15_DAYS:
            fill_one_row(
                workbook[sh.SHEET_EXPIRES_15.value],
                expiring_15_index,
                COLUMNS_LIST,
                user_info,
            )
            expiring_15_index += 1
        elif user_status == NTUserStatus.EXPIRING_30_DAYS:
            fill_one_row(
                workbook[sh.SHEET_EXPIRES_30.value],
                expiring_30_index,
                COLUMNS_LIST,
                user_info,
            )
            expiring_30_index += 1
        elif user_status == NTUserStatus.EXPIRING_60_DAYS:
            fill_one_row(
                workbook[sh.SHEET_EXPIRES_60.value],
                expiring_60_index,
                COLUMNS_LIST,
                user_info,
            )
            expiring_60_index += 1
    table_range = f"{shs.NAME_COLUMN.value}1:{shs.EXPIRATION_DATE_COLUMN.value}{all_users_index-1}"
    create_table_results(workbook[sh.SHEET_ALL_USERS.value], table_range)


def create_table_results(ws: Worksheet, tablerange: str):
    """Creates the table of results inside given worksheet for given tablerange.

    :param ws: The worksheet you want to create the table in.
    :param tablerange: The range of cells that will be in this table.
    """
    ws[f"{shs.NAME_COLUMN.value}1"] = tb.TABLE_NAME_COLUMN.value
    ws[f"{shs.NT_USER_COLUMN.value}1"] = tb.TABLE_NT_USER_COLUMN.value
    ws[f"{shs.EXPIRATION_DATE_COLUMN.value}1"] = tb.TABLE_EXPIRATION_DATE_COLUMN.value
    light_style = TableStyleInfo(name=tb.TABLE_STYLE.value, showRowStripes=True, showColumnStripes=True)
    if ws.tables:
        del ws.tables[tb.TABLE_NAME.value]
    table = Table(displayName=tb.TABLE_NAME.value, ref=tablerange)
    table.tableStyleInfo = light_style
    ws.add_table(table)
