import os

import pandas as pd
from openpyxl import Workbook, load_workbook

from nt_users_tool.constants import SHEET_INPUT
from nt_users_tool.excel_processing import create_results_sheets, fill_all_sheets, read_nt_users
from nt_users_tool.nt_user_info import NTUserInfo
from tests.constants import LIST_USERS_EXAMPLE


def test_read_nt_users():
    expected_value = LIST_USERS_EXAMPLE
    wb_input = load_workbook("tests/test_nt_users_name_example.xlsx", read_only=False)
    test_worksheet = wb_input[SHEET_INPUT]

    test_function_output = read_nt_users(test_worksheet)

    assert test_function_output == expected_value


def test_create_results_sheets():
    expected_workbook = load_workbook("tests/test_nt_users_sorted.xlsx", read_only=False)

    test_workbook = Workbook()
    create_results_sheets(test_workbook)
    del test_workbook["Sheet"]

    assert expected_workbook.sheetnames == test_workbook.sheetnames


def test_fill_all_sheets():
    # create the test workbook
    test_workbook = Workbook()
    create_results_sheets(test_workbook)
    list_nt_users = []
    for nt_users_names in LIST_USERS_EXAMPLE:
        list_nt_users.append(NTUserInfo("", nt_users_names, None))
    fill_all_sheets(test_workbook, list_nt_users)
    del test_workbook["Sheet"]
    test_workbook.save(filename="tests/test.xlsx")

    # use library pandas to compare the both workboot
    expected_workbook = pd.read_excel("tests/test_nt_users_sorted.xlsx")
    test_workbook_loaded = pd.read_excel("tests/test.xlsx")

    assert expected_workbook.equals(test_workbook_loaded)
    # remove the test workbook
    os.remove("tests/test.xlsx")
